#!/usr/bin/env python3
"""
Generate XLS documentation for Media Plan Schema

To generate documentation for a different schema version,
update the SCHEMA_VERSION variable below before running the script.
"""
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from pathlib import Path

# ============================================================
# CONFIGURATION: Update this to document a different version
# ============================================================
SCHEMA_VERSION = '3.0'  # Change to '1.0', '2.0', etc. as needed
# ============================================================


def get_enum_values(property_def):
    """Extract enum values from property definition"""
    if 'enum' in property_def:
        return ', '.join(str(v) for v in property_def['enum'])
    return ''


def get_data_type(property_def):
    """Extract data type from property definition"""
    prop_type = property_def.get('type', '')

    if prop_type == 'array':
        items = property_def.get('items', {})
        item_type = items.get('type', 'object')
        return f'array[{item_type}]'
    elif prop_type == 'string':
        format_type = property_def.get('format', '')
        if format_type:
            return f'string ({format_type})'

    return prop_type


def parse_schema_properties(schema_path, required_fields=None):
    """Parse a schema file and extract field information"""
    with open(schema_path, 'r') as f:
        schema = json.load(f)

    properties = schema.get('properties', {})
    required = required_fields if required_fields is not None else schema.get('required', [])

    fields = []
    for field_name, field_def in properties.items():
        # Skip $ref fields for now
        if '$ref' in field_def and 'description' not in field_def:
            continue

        fields.append({
            'name': field_name,
            'description': field_def.get('description', ''),
            'required': 'Yes' if field_name in required else 'No',
            'data_type': get_data_type(field_def),
            'enum': get_enum_values(field_def)
        })

    return fields, schema


def parse_meta_schema(schema_path):
    """Parse the main mediaplan schema and extract meta fields"""
    with open(schema_path, 'r') as f:
        schema = json.load(f)

    meta_def = schema['properties']['meta']
    properties = meta_def.get('properties', {})
    required = meta_def.get('required', [])

    fields = []
    for field_name, field_def in properties.items():
        fields.append({
            'name': field_name,
            'description': field_def.get('description', ''),
            'required': 'Yes' if field_name in required else 'No',
            'data_type': get_data_type(field_def),
            'enum': get_enum_values(field_def)
        })

    return fields


def parse_campaign_schema_with_objects(schema_path):
    """Parse campaign schema and extract fields including array item structures"""
    fields, schema = parse_schema_properties(schema_path)

    # Add target_audiences array item structure
    if 'target_audiences' in schema['properties']:
        target_audiences = schema['properties']['target_audiences']
        if target_audiences.get('type') == 'array' and 'items' in target_audiences:
            items_def = target_audiences['items']
            items_required = items_def.get('required', [])

            fields.append({
                'name': '',
                'description': '--- target_audiences item structure ---',
                'required': '',
                'data_type': '',
                'enum': ''
            })

            for prop_name, prop_def in items_def.get('properties', {}).items():
                fields.append({
                    'name': f'  {prop_name}',
                    'description': prop_def.get('description', ''),
                    'required': 'Yes' if prop_name in items_required else 'No',
                    'data_type': get_data_type(prop_def),
                    'enum': get_enum_values(prop_def)
                })

    # Add target_locations array item structure
    if 'target_locations' in schema['properties']:
        target_locations = schema['properties']['target_locations']
        if target_locations.get('type') == 'array' and 'items' in target_locations:
            items_def = target_locations['items']
            items_required = items_def.get('required', [])

            fields.append({
                'name': '',
                'description': '--- target_locations item structure ---',
                'required': '',
                'data_type': '',
                'enum': ''
            })

            for prop_name, prop_def in items_def.get('properties', {}).items():
                fields.append({
                    'name': f'  {prop_name}',
                    'description': prop_def.get('description', ''),
                    'required': 'Yes' if prop_name in items_required else 'No',
                    'data_type': get_data_type(prop_def),
                    'enum': get_enum_values(prop_def)
                })

    return fields


def parse_lineitem_schema_with_objects(schema_path):
    """Parse lineitem schema and extract fields including metric_formulas structure"""
    fields, schema = parse_schema_properties(schema_path)

    # Add metric_formulas object structure
    if 'metric_formulas' in schema['properties']:
        metric_formulas = schema['properties']['metric_formulas']
        if metric_formulas.get('type') == 'object' and 'additionalProperties' in metric_formulas:
            additional_props = metric_formulas['additionalProperties']

            fields.append({
                'name': '',
                'description': '--- metric_formulas object structure ---',
                'required': '',
                'data_type': '',
                'enum': ''
            })

            for prop_name, prop_def in additional_props.get('properties', {}).items():
                fields.append({
                    'name': f'  {prop_name}',
                    'description': prop_def.get('description', ''),
                    'required': 'No',
                    'data_type': get_data_type(prop_def),
                    'enum': get_enum_values(prop_def)
                })

    return fields


def parse_dictionary_schema(schema_path):
    """Parse the dictionary schema and flatten it for documentation"""
    with open(schema_path, 'r') as f:
        schema = json.load(f)

    fields = []

    # Get all $defs for reference
    defs = schema.get('$defs', {})
    custom_field_def = defs.get('custom_field_config', {})
    metric_formula_def = defs.get('metric_formula_config', {})
    custom_metric_def = defs.get('custom_metric_config', {})

    config_required = custom_field_def.get('required', [])
    custom_metric_required = custom_metric_def.get('required', [])

    # Add top-level groups
    for group_name, group_def in schema['properties'].items():
        fields.append({
            'name': group_name,
            'description': group_def.get('description', ''),
            'required': 'No',
            'data_type': 'object',
            'enum': ''
        })

        # Determine the correct object type based on the group
        if group_name in ['meta_custom_dimensions', 'campaign_custom_dimensions', 'lineitem_custom_dimensions']:
            object_type = 'custom_field_config'
        elif group_name == 'standard_metrics':
            object_type = 'metric_formula_config'
        elif group_name == 'custom_metrics':
            object_type = 'custom_metric_config'
        elif group_name == 'custom_costs':
            object_type = 'custom_field_config'
        else:
            object_type = 'object'

        # Add individual custom fields within the group
        for field_name in group_def['properties'].keys():
            fields.append({
                'name': f'  {field_name}',
                'description': f'Configuration for {field_name}',
                'required': 'No',
                'data_type': f'object ({object_type})',
                'enum': ''
            })

    # Add custom_field_config properties as a reference section
    if custom_field_def:
        fields.append({
            'name': '',
            'description': '--- custom_field_config structure ---',
            'required': '',
            'data_type': '',
            'enum': ''
        })

        for prop_name, prop_def in custom_field_def['properties'].items():
            fields.append({
                'name': f'  {prop_name}',
                'description': prop_def.get('description', ''),
                'required': 'Yes' if prop_name in config_required else 'Conditional',
                'data_type': get_data_type(prop_def),
                'enum': get_enum_values(prop_def)
            })

    # Add metric_formula_config properties as a reference section
    if metric_formula_def:
        fields.append({
            'name': '',
            'description': '--- metric_formula_config structure ---',
            'required': '',
            'data_type': '',
            'enum': ''
        })

        for prop_name, prop_def in metric_formula_def['properties'].items():
            fields.append({
                'name': f'  {prop_name}',
                'description': prop_def.get('description', ''),
                'required': 'No',
                'data_type': get_data_type(prop_def),
                'enum': get_enum_values(prop_def)
            })

    # Add custom_metric_config properties as a reference section
    if custom_metric_def:
        fields.append({
            'name': '',
            'description': '--- custom_metric_config structure ---',
            'required': '',
            'data_type': '',
            'enum': ''
        })

        for prop_name, prop_def in custom_metric_def['properties'].items():
            fields.append({
                'name': f'  {prop_name}',
                'description': prop_def.get('description', ''),
                'required': 'Yes' if prop_name in custom_metric_required else 'Conditional',
                'data_type': get_data_type(prop_def),
                'enum': get_enum_values(prop_def)
            })

    return fields


def format_worksheet(ws):
    """Apply formatting to worksheet"""
    # Header row formatting
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

    # Set column widths (increased description column by 50%)
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 90
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 30

    # Format data rows
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

    # Freeze header row
    ws.freeze_panes = 'A2'


def write_fields_to_sheet(ws, fields):
    """Write field data to worksheet"""
    # Headers
    ws.append(['Name', 'Description', 'Required', 'Data Type', 'Enum'])

    # Data
    for field in fields:
        ws.append([
            field['name'],
            field['description'],
            field['required'],
            field['data_type'],
            field['enum']
        ])

    format_worksheet(ws)


def main():
    # Get project root (parent of scripts directory)
    project_root = Path(__file__).parent.parent
    schema_dir = project_root / 'schemas' / SCHEMA_VERSION

    print(f"Generating documentation for schema version {SCHEMA_VERSION}")
    print(f"Schema directory: {schema_dir}\n")

    # Validate schema directory exists
    if not schema_dir.exists():
        print(f"Error: Schema directory not found: {schema_dir}")
        return

    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Meta worksheet
    print("Processing meta schema...")
    ws_meta = wb.create_sheet('meta')
    meta_fields = parse_meta_schema(schema_dir / 'mediaplan.schema.json')
    write_fields_to_sheet(ws_meta, meta_fields)

    # Campaign worksheet
    print("Processing campaign schema...")
    ws_campaign = wb.create_sheet('campaign')
    campaign_fields = parse_campaign_schema_with_objects(schema_dir / 'campaign.schema.json')
    write_fields_to_sheet(ws_campaign, campaign_fields)

    # Line Items worksheet
    print("Processing lineitem schema...")
    ws_lineitems = wb.create_sheet('lineitems')
    lineitem_fields = parse_lineitem_schema_with_objects(schema_dir / 'lineitem.schema.json')
    write_fields_to_sheet(ws_lineitems, lineitem_fields)

    # Dictionary worksheet (optional - only in v2.0+)
    dictionary_fields = []
    dictionary_schema_path = schema_dir / 'dictionary.schema.json'
    if dictionary_schema_path.exists():
        print("Processing dictionary schema...")
        ws_dictionary = wb.create_sheet('dictionary')
        dictionary_fields = parse_dictionary_schema(dictionary_schema_path)
        write_fields_to_sheet(ws_dictionary, dictionary_fields)
    else:
        print("Skipping dictionary schema (not present in this version)...")

    # Save workbook to the schema version directory
    output_file = schema_dir / f'mediaplan_schema_v{SCHEMA_VERSION}_documentation.xlsx'
    wb.save(output_file)
    print(f"\nSchema documentation generated: {output_file}")

    # Summary
    summary = f"Worksheets: meta ({len(meta_fields)} fields), campaign ({len(campaign_fields)} fields), " \
              f"lineitems ({len(lineitem_fields)} fields)"
    if dictionary_fields:
        summary += f", dictionary ({len(dictionary_fields)} rows)"
    print(summary)


if __name__ == '__main__':
    main()
